from __future__ import annotations

import csv
import io
import os
import subprocess
import sys
import threading
import wsgiref.simple_server
import xml.etree.ElementTree as ET
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Generic, TypeVar
from urllib.parse import parse_qs, urlparse

import requests

from .excel_xml import _classify_rows, _effective_max_columns, _infer_headers, _trim_trailing_empty_cells, parse_workbook_bytes
from .models import CellData, RowData, SheetData, WorkbookData

try:
    from openpyxl import load_workbook as load_xlsx_workbook
except Exception:  # pragma: no cover - optional dependency path
    load_xlsx_workbook = None

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
except Exception:  # pragma: no cover - optional dependency path
    service_account = None
    build = None
    MediaIoBaseDownload = None

try:
    from google.auth.transport.requests import Request as GoogleAuthRequest
except Exception:  # pragma: no cover - optional dependency path
    GoogleAuthRequest = None

try:
    from google.oauth2.credentials import Credentials as UserCredentials
except Exception:  # pragma: no cover - optional dependency path
    UserCredentials = None

try:
    import google_auth_oauthlib.flow as oauth_flow_module
    from google_auth_oauthlib.flow import InstalledAppFlow
except Exception:  # pragma: no cover - optional dependency path
    oauth_flow_module = None
    InstalledAppFlow = None


SOURCE_LOCAL = "local"
SOURCE_SVN = "svn"
SOURCE_GOOGLE_SHEETS = "google_sheets"
GOOGLE_SHEETS_READONLY_SCOPE = "https://www.googleapis.com/auth/spreadsheets.readonly"
GOOGLE_DRIVE_READONLY_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
GOOGLE_READONLY_SCOPES = [GOOGLE_SHEETS_READONLY_SCOPE, GOOGLE_DRIVE_READONLY_SCOPE]
LOCAL_TABLE_EXTENSIONS = {".xml", ".xlsx", ".csv"}


@dataclass(frozen=True, slots=True)
class WorkbookSource:
    kind: str
    file_path: str = ""
    revision: str = ""
    display_name: str = ""
    source_root: str = ""
    metadata: tuple[tuple[str, str], ...] = ()

    @property
    def metadata_dict(self) -> dict[str, str]:
        return dict(self.metadata)


@dataclass(slots=True)
class RevisionEntry:
    revision: str
    author: str
    date: str
    message: str


@dataclass(frozen=True, slots=True)
class SourceFileEntry:
    name: str
    path: str


@dataclass(frozen=True, slots=True)
class SourceBrowseEntry:
    name: str
    path: str
    kind: str
    revision: str = ""
    author: str = ""
    date: str = ""
    size: str = ""


_K = TypeVar("_K")
_V = TypeVar("_V")


class _BoundedCache(Generic[_K, _V]):
    __slots__ = ("_capacity", "_store", "_lock")

    def __init__(self, capacity: int) -> None:
        self._capacity = max(1, int(capacity))
        self._store: "OrderedDict[_K, _V]" = OrderedDict()
        self._lock = threading.Lock()

    def get(self, key: _K, default: _V | None = None) -> _V | None:
        with self._lock:
            try:
                value = self._store[key]
            except KeyError:
                return default
            self._store.move_to_end(key)
            return value

    def __contains__(self, key: _K) -> bool:
        with self._lock:
            return key in self._store

    def __getitem__(self, key: _K) -> _V:
        with self._lock:
            value = self._store[key]
            self._store.move_to_end(key)
            return value

    def __setitem__(self, key: _K, value: _V) -> None:
        with self._lock:
            if key in self._store:
                self._store.move_to_end(key)
            self._store[key] = value
            while len(self._store) > self._capacity:
                self._store.popitem(last=False)

    def pop(self, key: _K, default: _V | None = None) -> _V | None:
        with self._lock:
            return self._store.pop(key, default)

    def clear(self) -> None:
        with self._lock:
            self._store.clear()

    def __len__(self) -> int:
        with self._lock:
            return len(self._store)


_WORKBOOK_CACHE: _BoundedCache[str, WorkbookData] = _BoundedCache(32)
_REVISION_CACHE: _BoundedCache[str, list[RevisionEntry]] = _BoundedCache(64)
_GOOGLE_SHEETS_SERVICE = None
_GOOGLE_DRIVE_SERVICE = None
_GOOGLE_METADATA_CACHE: _BoundedCache[str, dict] = _BoundedCache(64)
_GOOGLE_AUTH_MODE_OVERRIDE = "service_account"
_GOOGLE_SERVICE_ACCOUNT_OVERRIDE = ""
_GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE = ""
_GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE = ""


def load_workbook_from_source(source: WorkbookSource) -> WorkbookData:
    effective_source = source
    if source.kind == SOURCE_SVN:
        resolved_revision = _resolve_svn_revision(
            normalize_source_target(source.file_path),
            source.revision or "HEAD",
        )
        if resolved_revision and resolved_revision != (source.revision or ""):
            effective_source = WorkbookSource(
                kind=source.kind,
                file_path=source.file_path,
                revision=resolved_revision,
                display_name=source.display_name,
                source_root=source.source_root,
                metadata=source.metadata,
            )

    snapshot_id = build_snapshot_id(effective_source)
    cached = _WORKBOOK_CACHE.get(snapshot_id)
    if cached is not None:
        return cached

    if effective_source.kind == SOURCE_LOCAL:
        file_path = Path(effective_source.file_path)
        suffix = file_path.suffix.lower()
        if suffix == ".xml":
            xml_bytes = file_path.read_bytes()
            workbook = parse_workbook_bytes(
                xml_bytes,
                path=file_path,
                source_kind=SOURCE_LOCAL,
                source_label=effective_source.display_name or file_path.name,
                snapshot_id=snapshot_id,
                source_meta={**effective_source.metadata_dict, "file_format": "xml"},
            )
        elif suffix == ".xlsx":
            workbook = load_xlsx_local_workbook(
                file_path, snapshot_id=snapshot_id, source_meta=effective_source.metadata_dict
            )
        elif suffix == ".csv":
            workbook = load_csv_local_workbook(
                file_path, snapshot_id=snapshot_id, source_meta=effective_source.metadata_dict
            )
        else:
            raise ValueError(f"不支持的本地文件类型: {file_path.suffix}")
    elif effective_source.kind == SOURCE_SVN:
        target = normalize_source_target(effective_source.file_path)
        revision = effective_source.revision or "HEAD"
        xml_bytes = svn_cat_bytes(target, revision)
        workbook = parse_workbook_bytes(
            xml_bytes,
            path=None,
            source_kind=SOURCE_SVN,
            source_label=effective_source.display_name or f"{source_path_name(target)}@{revision}",
            snapshot_id=snapshot_id,
            source_meta={**effective_source.metadata_dict, "revision": revision, "source_path": target},
        )
    elif effective_source.kind == SOURCE_GOOGLE_SHEETS:
        workbook = load_google_workbook(effective_source)
    else:
        raise ValueError(f"未知来源类型: {effective_source.kind}")

    _WORKBOOK_CACHE[snapshot_id] = workbook
    return workbook


def _resolve_svn_revision(target: str, revision: str) -> str:
    value = (revision or "").strip()
    if value and value.upper() != "HEAD":
        return value
    command = ["svn", "info", "--show-item", "revision", "-r", "HEAD", target]
    try:
        process = run_hidden_process(command)
    except FileNotFoundError:
        return value or "HEAD"
    if process.returncode != 0:
        return value or "HEAD"
    resolved = decode_text_bytes(process.stdout).strip()
    return resolved or (value or "HEAD")


def set_google_auth_settings(
    auth_mode: str,
    service_account_path: str = "",
    oauth_client_path: str = "",
    oauth_token_path: str = "",
) -> None:
    global _GOOGLE_AUTH_MODE_OVERRIDE
    global _GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE
    global _GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE
    global _GOOGLE_SERVICE_ACCOUNT_OVERRIDE
    normalized_mode = _normalize_google_auth_mode(auth_mode)
    normalized_service_path = str(service_account_path or "").strip()
    normalized_client_path = str(oauth_client_path or "").strip()
    normalized_token_path = str(oauth_token_path or "").strip()
    if (
        normalized_mode == _GOOGLE_AUTH_MODE_OVERRIDE
        and normalized_service_path == _GOOGLE_SERVICE_ACCOUNT_OVERRIDE
        and normalized_client_path == _GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE
        and normalized_token_path == _GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE
    ):
        return
    _GOOGLE_AUTH_MODE_OVERRIDE = normalized_mode
    _GOOGLE_SERVICE_ACCOUNT_OVERRIDE = normalized_service_path
    _GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE = normalized_client_path
    _GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE = normalized_token_path
    _reset_google_service_cache()


def set_google_service_account_path(path: str) -> None:
    set_google_auth_settings(
        _GOOGLE_AUTH_MODE_OVERRIDE,
        service_account_path=path,
        oauth_client_path=_GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE,
        oauth_token_path=_GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE,
    )


def default_google_oauth_token_path() -> str:
    localappdata = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA")
    if localappdata:
        return str(Path(localappdata) / "FenJiuBiHe" / "google_oauth_token.json")
    return str(Path.cwd() / "google_oauth_token.json")


def start_google_oauth_login(
    client_json_path: str,
    token_json_path: str = "",
    manual_prompt_callback: Callable[[str], bool] | None = None,
) -> str:
    client_value = str(client_json_path or "").strip()
    if not client_value:
        raise RuntimeError("请先在“管理快捷配置”中选择 OAuth 客户端 credentials.json。")
    client_path = Path(client_value)
    if not client_path.is_file():
        raise RuntimeError("OAuth 客户端 JSON 无效，请确认你选择的是一个实际存在的 credentials.json 文件。")
    if InstalledAppFlow is None:
        raise RuntimeError("当前环境缺少 google-auth-oauthlib，无法执行 Google 登录。")
    token_path = Path(str(token_json_path or "").strip() or default_google_oauth_token_path())
    flow = InstalledAppFlow.from_client_secrets_file(str(client_path), GOOGLE_READONLY_SCOPES)
    if manual_prompt_callback is not None:
        credentials = _run_google_oauth_login_manual(flow, manual_prompt_callback)
        _save_google_oauth_token(credentials, token_path)
        return str(token_path)
    try:
        credentials = flow.run_local_server(port=0, open_browser=True)
    except PermissionError as exc:
        credentials = _run_google_oauth_login_manual(flow, manual_prompt_callback)
    except OSError as exc:
        if str(exc).strip().endswith(": ''"):
            credentials = _run_google_oauth_login_manual(flow, manual_prompt_callback)
        else:
            raise
    _save_google_oauth_token(credentials, token_path)
    return str(token_path)


def build_snapshot_id(source: WorkbookSource) -> str:
    if source.kind == SOURCE_LOCAL:
        file_path = Path(source.file_path)
        stat = file_path.stat()
        return f"{SOURCE_LOCAL}:{file_path.resolve()}:{stat.st_mtime_ns}:{stat.st_size}"
    if source.kind == SOURCE_SVN:
        return f"{SOURCE_SVN}:{normalize_source_target(source.file_path)}:{source.revision or 'HEAD'}"
    if source.kind == SOURCE_GOOGLE_SHEETS:
        parsed = parse_google_sheet_target(source.file_path)
        return f"{SOURCE_GOOGLE_SHEETS}:{parsed['spreadsheet_id']}:{parsed['gid']}"
    return f"{source.kind}:{source.file_path}:{source.revision}"


def describe_google_sheet(value: str | Path) -> dict:
    parsed = parse_google_sheet_target(value)
    spreadsheet_id = parsed["spreadsheet_id"]
    if not spreadsheet_id:
        raise RuntimeError("Google Sheets 链接或 ID 无效。")

    metadata_error: Exception | None = None
    try:
        metadata = _fetch_google_sheet_metadata(spreadsheet_id)
        sheet_summaries = [
            {
                "sheet_id": str(sheet.get("properties", {}).get("sheetId", "")),
                "title": str(sheet.get("properties", {}).get("title", "")),
                "row_count": int(sheet.get("properties", {}).get("gridProperties", {}).get("rowCount", 0) or 0),
                "column_count": int(sheet.get("properties", {}).get("gridProperties", {}).get("columnCount", 0) or 0),
            }
            for sheet in metadata.get("sheets", [])
        ]
        if sheet_summaries:
            return {
                "spreadsheet_id": spreadsheet_id,
                "gid": parsed["gid"],
                "title": str(metadata.get("properties", {}).get("title") or spreadsheet_id),
                "sheets": sheet_summaries,
            }
    except Exception as exc:  # noqa: BLE001
        metadata_error = exc

    try:
        exported = _download_google_sheet_as_xlsx_bytes(spreadsheet_id)
        return _describe_google_sheet_xlsx_export(
            spreadsheet_id,
            parsed["gid"],
            title=str(exported["title"]),
            content=bytes(exported["content"]),
        )
    except Exception as export_error:  # noqa: BLE001
        if parsed["gid"]:
            return {
                "spreadsheet_id": spreadsheet_id,
                "gid": parsed["gid"],
                "title": spreadsheet_id,
                "sheets": [
                    {
                        "sheet_id": parsed["gid"],
                        "title": f"gid_{parsed['gid']}",
                        "row_count": 0,
                        "column_count": 0,
                    }
                ],
            }
        if metadata_error is not None:
            raise RuntimeError(
                "无法读取 Google Sheets 元数据。\n"
                f"Sheets API 失败：{metadata_error}\n"
                f"Drive 导出回退失败：{export_error}"
            ) from export_error
        raise


def _describe_google_sheet_xlsx_export(
    spreadsheet_id: str,
    gid: str,
    *,
    title: str,
    content: bytes,
) -> dict[str, object]:
    if load_xlsx_workbook is None:
        raise RuntimeError("当前环境缺少 openpyxl，无法读取 Google 导出的 xlsx。")
    xlsx = load_xlsx_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_summaries = [
            {
                "sheet_id": "",
                "title": ws.title,
                "row_count": int(ws.max_row or 0),
                "column_count": int(ws.max_column or 0),
            }
            for ws in xlsx.worksheets
        ]
    finally:
        xlsx.close()
    return {
        "spreadsheet_id": spreadsheet_id,
        "gid": gid,
        "title": title or spreadsheet_id,
        "sheets": sheet_summaries,
    }


def list_local_table_files(root: str | Path) -> list[SourceFileEntry]:
    base = Path(root)
    if base.is_file():
        if base.suffix.lower() in LOCAL_TABLE_EXTENSIONS:
            return [SourceFileEntry(name=base.name, path=str(base))]
        return []
    scan_dirs: list[Path] = []
    for candidate in [*preferred_xml_roots(root), str(base)]:
        directory = Path(candidate)
        if directory.exists() and directory.is_dir():
            scan_dirs.append(directory)
    seen_paths: set[str] = set()
    files: list[SourceFileEntry] = []
    for directory in _dedupe_paths(scan_dirs):
        for path in sorted(directory.iterdir()):
            if not path.is_file() or path.suffix.lower() not in LOCAL_TABLE_EXTENSIONS:
                continue
            normalized = str(path)
            if normalized in seen_paths:
                continue
            seen_paths.add(normalized)
            files.append(SourceFileEntry(name=path.name, path=normalized))
    return sorted(files, key=lambda item: item.path.lower())


def list_svn_xml_files(root: str, recursive: bool = True, revision: str = "HEAD") -> list[SourceFileEntry]:
    target = normalize_source_target(root)
    revision = str(revision or "HEAD").strip() or "HEAD"
    attempted: list[str] = []
    aggregated: dict[str, SourceFileEntry] = {}
    for candidate in preferred_xml_roots(target):
        attempted.append(candidate)
        try:
            for item in _list_svn_xml_files_from_target(candidate, recursive=recursive, revision=revision):
                aggregated[item.path] = item
        except RuntimeError:
            continue
    if aggregated:
        return sorted(aggregated.values(), key=lambda item: item.path.lower())
    return _list_svn_xml_files_from_target(target, recursive=recursive, revision=revision)


def list_svn_directory(root: str | Path) -> list[SourceBrowseEntry]:
    target = normalize_source_target(root)
    output = run_svn_command(["svn", "list", "--xml", target])
    try:
        node = ET.fromstring(output)
    except ET.ParseError as exc:
        raise RuntimeError(f"解析 SVN 目录失败: {exc}") from exc

    entries: list[SourceBrowseEntry] = []
    for entry_node in node.findall(".//entry"):
        kind = (entry_node.attrib.get("kind") or "").strip() or "file"
        name = (entry_node.findtext("name") or "").strip()
        if not name:
            continue
        commit_node = entry_node.find("commit")
        revision = ""
        author = ""
        date = ""
        if commit_node is not None:
            revision = (commit_node.attrib.get("revision") or "").strip()
            author = (commit_node.findtext("author") or "").strip()
            date = _display_log_date((commit_node.findtext("date") or "").strip())
        size = (entry_node.findtext("size") or "").strip()
        entries.append(
            SourceBrowseEntry(
                name=name,
                path=join_source_target(target, name),
                kind=kind,
                revision=revision,
                author=author,
                date=date,
                size=size,
            )
        )
    entries.sort(key=lambda item: (item.kind != "dir", item.name.lower()))
    return entries


def list_svn_revisions(file_path: str | Path, limit: int = 50) -> list[RevisionEntry]:
    target = normalize_source_target(file_path)
    cache_key = f"{target}|{limit}"
    cached = _REVISION_CACHE.get(cache_key)
    if cached is not None:
        return cached

    output = run_svn_command(["svn", "log", "--xml", "-l", str(limit), target])
    try:
        root = ET.fromstring(output)
    except ET.ParseError as exc:
        raise RuntimeError(f"解析 SVN 日志失败: {exc}") from exc

    rows: list[RevisionEntry] = [RevisionEntry("HEAD", "-", "", "最新版本")]
    for node in root.findall(".//logentry"):
        revision = (node.attrib.get("revision") or "").strip()
        author = (node.findtext("author") or "").strip() or "-"
        date = _display_log_date((node.findtext("date") or "").strip())
        message = (node.findtext("msg") or "").strip() or "(无提交说明)"
        if revision:
            rows.append(RevisionEntry(revision=revision, author=author, date=date, message=message))
    _REVISION_CACHE[cache_key] = rows
    return rows


def svn_cat_bytes(file_path: str | Path, revision: str) -> bytes:
    command = ["svn", "cat", "-r", revision or "HEAD", normalize_source_target(file_path)]
    process = run_hidden_process(command)
    if process.returncode != 0:
        details = decode_text_bytes(process.stderr).strip() or decode_text_bytes(process.stdout).strip()
        raise RuntimeError(f"读取 SVN 版本内容失败: {' '.join(command)}\n{details}")
    return process.stdout


def verify_svn_file(file_path: str | Path) -> tuple[bool, str]:
    command = ["svn", "info", normalize_source_target(file_path)]
    process = run_hidden_process(command)
    if process.returncode == 0:
        return True, ""
    details = decode_text_bytes(process.stderr).strip() or decode_text_bytes(process.stdout).strip()
    return False, details


def run_svn_command(args: list[str]) -> str:
    process = run_hidden_process(args)
    stdout = decode_text_bytes(process.stdout)
    stderr = decode_text_bytes(process.stderr)
    if process.returncode != 0:
        details = stderr.strip() or stdout.strip()
        raise RuntimeError(f"SVN 命令执行失败: {' '.join(args)}\n{details}")
    return stdout


def run_hidden_process(args: list[str]) -> subprocess.CompletedProcess[bytes]:
    kwargs = {
        "capture_output": True,
        "text": False,
        "check": False,
    }
    if sys.platform.startswith("win"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0
        kwargs["startupinfo"] = startupinfo
        kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    return subprocess.run(args, **kwargs)


def decode_text_bytes(data: bytes) -> str:
    if not data:
        return ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "cp936"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _display_log_date(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    return text.replace("T", " ").replace("Z", "")[:19]


def normalize_source_target(value: str | Path) -> str:
    text = str(value or "").strip()
    if text.lower().startswith("svn:") and not text.lower().startswith("svn://"):
        return f"svn://{text[4:].lstrip('/\\').replace('\\', '/')}"
    if _looks_like_url(text):
        return text.replace("\\", "/")
    return str(Path(text))


def join_source_target(root: str | Path, child: str) -> str:
    base = normalize_source_target(root)
    item = str(child or "").strip().replace("\\", "/")
    if not item:
        return base
    if _looks_like_url(item):
        return normalize_source_target(item)
    if not _looks_like_url(base):
        child_path = Path(item)
        if child_path.is_absolute():
            return str(child_path)
    if _looks_like_url(base):
        return f"{base.rstrip('/')}/{item.lstrip('/')}"
    return str(Path(base).joinpath(*[part for part in item.split("/") if part]))


def source_path_name(value: str | Path) -> str:
    text = str(value or "").strip()
    if is_google_sheets_target(text):
        parsed = parse_google_sheet_target(text)
        if parsed["gid"]:
            return f"{parsed['spreadsheet_id']}#gid={parsed['gid']}"
        return parsed["spreadsheet_id"]
    text = normalize_source_target(value).rstrip("/")
    if not text:
        return ""
    return text.split("/")[-1].split("\\")[-1]


def source_relative_path(root: str | Path, value: str | Path) -> str:
    base = normalize_source_target(root).replace("\\", "/").rstrip("/")
    target = normalize_source_target(value).replace("\\", "/").rstrip("/")
    if not base or not target:
        return source_path_name(value)
    if target.lower().startswith(f"{base.lower()}/"):
        return target[len(base) + 1 :]
    try:
        return str(Path(target).relative_to(Path(base))).replace("\\", "/")
    except (ValueError, RuntimeError):
        return source_path_name(value)


def infer_source_kind(value: str | Path) -> str:
    text = str(value or "").strip()
    if not text:
        return SOURCE_LOCAL
    if is_google_sheets_target(text):
        return SOURCE_GOOGLE_SHEETS
    return SOURCE_SVN if _looks_like_url(text) else SOURCE_LOCAL


def parse_google_sheet_target(value: str | Path) -> dict[str, str]:
    text = str(value or "").strip()
    if not text:
        return {"spreadsheet_id": "", "gid": ""}
    lowered = text.lower()
    if "docs.google.com/spreadsheets/d/" in lowered:
        parsed = urlparse(text)
        parts = [part for part in parsed.path.split("/") if part]
        spreadsheet_id = ""
        for index, part in enumerate(parts):
            if part == "d" and index + 1 < len(parts):
                spreadsheet_id = parts[index + 1]
                break
        query = parse_qs(parsed.query)
        gid = (query.get("gid") or [""])[0]
        if not gid and parsed.fragment.startswith("gid="):
            gid = parsed.fragment[4:]
        return {"spreadsheet_id": spreadsheet_id, "gid": gid}
    return {"spreadsheet_id": text, "gid": ""}


def is_google_sheets_target(value: str | Path) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    lowered = text.lower()
    if "docs.google.com/spreadsheets/d/" in lowered:
        return True
    if any(token in text for token in ("\\", "/", ":")):
        return False
    return len(text) >= 30 and all(ch.isalnum() or ch in "-_" for ch in text)


def load_google_workbook(source: WorkbookSource) -> WorkbookData:
    parsed = parse_google_sheet_target(source.file_path)
    spreadsheet_id = parsed["spreadsheet_id"]
    if not spreadsheet_id:
        raise RuntimeError("Google Sheets 链接或 ID 无效。")

    drive_error: Exception | None = None
    try:
        exported = _download_google_sheet_as_xlsx_bytes(spreadsheet_id)
        return load_xlsx_bytes_workbook(
            exported["content"],
            file_name=str(exported["file_name"]),
            snapshot_id=build_snapshot_id(source),
            source_kind=SOURCE_GOOGLE_SHEETS,
            source_label=str(exported["title"]),
            source_meta={
                **source.metadata_dict,
                "spreadsheet_id": spreadsheet_id,
                "gid": parsed["gid"],
                "source_path": str(source.file_path),
                "google_load_mode": "drive_export",
            },
        )
    except Exception as exc:  # noqa: BLE001
        drive_error = exc

    try:
        description = describe_google_sheet(source.file_path)
        sheet_names = [item["title"] for item in description["sheets"] if item["title"]]
        if not sheet_names:
            raise RuntimeError("该 Google Sheets 没有可读取的工作表。")
        workbook = WorkbookData(
            path=None,
            tree=None,
            sheet_names=sheet_names,
            xml_bytes=b"",
            source_kind=SOURCE_GOOGLE_SHEETS,
            source_label=description["title"],
            snapshot_id=build_snapshot_id(source),
            source_meta={
                **source.metadata_dict,
                "spreadsheet_id": spreadsheet_id,
                "gid": description["gid"],
                "source_path": str(source.file_path),
                "google_load_mode": "sheets_api",
            },
        )
        sheet_gid_map = {item["title"]: item["sheet_id"] for item in description["sheets"] if item["title"]}
        workbook._sheet_loader = lambda sheet_name: _load_google_sheet_as_sheetdata(
            spreadsheet_id,
            sheet_name,
            gid=sheet_gid_map.get(sheet_name, ""),
        )
        return workbook
    except Exception as fallback_exc:  # noqa: BLE001
        if drive_error is None:
            raise
        raise RuntimeError(
            "无法读取 Google Sheets。\n"
            f"Drive 导出失败：{drive_error}\n"
            f"Sheets API 回退失败：{fallback_exc}"
        ) from fallback_exc


def load_xlsx_local_workbook(path: Path, *, snapshot_id: str, source_meta: dict[str, str] | None = None) -> WorkbookData:
    if load_xlsx_workbook is None:
        raise RuntimeError("当前环境缺少 openpyxl，无法读取 .xlsx 文件。")
    xlsx = load_xlsx_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_names = list(xlsx.sheetnames)
        sheet_cache: dict[str, SheetData] = {}
        for sheet_name in sheet_names:
            sheet = _load_xlsx_sheet_as_sheetdata(xlsx, sheet_name)
            if sheet is not None:
                sheet_cache[sheet_name] = sheet
    finally:
        xlsx.close()
    workbook = WorkbookData(
        path=path,
        tree=None,
        sheet_names=sheet_names,
        xml_bytes=b"",
        source_kind=SOURCE_LOCAL,
        source_label=path.name,
        snapshot_id=snapshot_id,
        source_meta={**(source_meta or {}), "file_format": "xlsx"},
    )
    workbook._sheet_cache = sheet_cache
    workbook._sheet_loader = lambda sheet_name: _load_xlsx_sheet_as_sheetdata_from_path(path, sheet_name)
    return workbook


def load_xlsx_bytes_workbook(
    content: bytes,
    *,
    file_name: str,
    snapshot_id: str,
    source_kind: str,
    source_label: str,
    source_meta: dict[str, str] | None = None,
) -> WorkbookData:
    if load_xlsx_workbook is None:
        raise RuntimeError("当前环境缺少 openpyxl，无法读取 .xlsx 文件。")
    xlsx = load_xlsx_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_names = list(xlsx.sheetnames)
        sheet_cache: dict[str, SheetData] = {}
        for sheet_name in sheet_names:
            sheet = _load_xlsx_sheet_as_sheetdata(xlsx, sheet_name)
            if sheet is not None:
                sheet_cache[sheet_name] = sheet
    finally:
        xlsx.close()
    workbook = WorkbookData(
        path=None,
        tree=None,
        sheet_names=sheet_names,
        xml_bytes=b"",
        source_kind=source_kind,
        source_label=source_label or file_name,
        snapshot_id=snapshot_id,
        source_meta={**(source_meta or {}), "file_format": "xlsx"},
    )
    workbook._sheet_cache = sheet_cache
    workbook._sheet_loader = lambda sheet_name: _load_xlsx_sheet_as_sheetdata_from_bytes(content, sheet_name)
    return workbook


def load_csv_local_workbook(path: Path, *, snapshot_id: str, source_meta: dict[str, str] | None = None) -> WorkbookData:
    workbook = WorkbookData(
        path=path,
        tree=None,
        sheet_names=[path.stem or path.name],
        xml_bytes=b"",
        source_kind=SOURCE_LOCAL,
        source_label=path.name,
        snapshot_id=snapshot_id,
        source_meta={**(source_meta or {}), "file_format": "csv"},
    )
    workbook._sheet_loader = lambda sheet_name: _load_csv_sheet_as_sheetdata(path, sheet_name)
    return workbook


def _load_xlsx_sheet_as_sheetdata(xlsx, sheet_name: str) -> SheetData | None:
    if sheet_name not in xlsx.sheetnames:
        return None
    worksheet = xlsx[sheet_name]
    rows: list[RowData] = []
    max_columns = 0
    for row_index, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        row = RowData(row_index=row_index)
        last_non_empty = 0
        for column_index, value in enumerate(raw_row, start=1):
            text = _stringify_tabular_value(value)
            if text:
                row.cells.append(CellData(column_index=column_index, value=text, data_type="String"))
                last_non_empty = column_index
        max_columns = max(max_columns, last_non_empty)
        rows.append(row)
    _trim_trailing_empty_cells(rows)
    max_columns = min(max_columns, _effective_max_columns(rows))
    sheet = SheetData(name=sheet_name, rows=rows, max_columns=max_columns)
    _infer_headers(sheet)
    _classify_rows(sheet)
    return sheet


def _load_xlsx_sheet_as_sheetdata_from_path(path: Path, sheet_name: str) -> SheetData | None:
    if load_xlsx_workbook is None:
        raise RuntimeError("当前环境缺少 openpyxl，无法读取 .xlsx 文件。")
    xlsx = load_xlsx_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return _load_xlsx_sheet_as_sheetdata(xlsx, sheet_name)
    finally:
        xlsx.close()


def _load_xlsx_sheet_as_sheetdata_from_bytes(content: bytes, sheet_name: str) -> SheetData | None:
    if load_xlsx_workbook is None:
        raise RuntimeError("当前环境缺少 openpyxl，无法读取 .xlsx 文件。")
    xlsx = load_xlsx_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    try:
        return _load_xlsx_sheet_as_sheetdata(xlsx, sheet_name)
    finally:
        xlsx.close()


def _load_csv_sheet_as_sheetdata(path: Path, sheet_name: str) -> SheetData:
    rows: list[RowData] = []
    max_columns = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row_index, raw_row in enumerate(csv.reader(handle), start=1):
            row = RowData(row_index=row_index)
            last_non_empty = 0
            for column_index, value in enumerate(raw_row, start=1):
                text = str(value or "")
                if text:
                    row.cells.append(CellData(column_index=column_index, value=text, data_type="String"))
                    last_non_empty = column_index
            max_columns = max(max_columns, last_non_empty)
            rows.append(row)
    _trim_trailing_empty_cells(rows)
    max_columns = min(max_columns, _effective_max_columns(rows))
    sheet = SheetData(name=sheet_name, rows=rows, max_columns=max_columns)
    _infer_headers(sheet)
    _classify_rows(sheet)
    return sheet


def _load_google_sheet_as_sheetdata(spreadsheet_id: str, sheet_name: str, gid: str = "") -> SheetData:
    values = _fetch_google_sheet_values(spreadsheet_id, sheet_name, gid=gid)
    rows: list[RowData] = []
    max_columns = 0
    for row_index, raw_row in enumerate(values, start=1):
        row = RowData(row_index=row_index)
        last_non_empty = 0
        for column_index, value in enumerate(raw_row, start=1):
            text = _stringify_tabular_value(value)
            if text:
                row.cells.append(CellData(column_index=column_index, value=text, data_type="String"))
                last_non_empty = column_index
        max_columns = max(max_columns, last_non_empty)
        rows.append(row)
    _trim_trailing_empty_cells(rows)
    max_columns = min(max_columns, _effective_max_columns(rows))
    sheet = SheetData(name=sheet_name, rows=rows, max_columns=max_columns)
    _infer_headers(sheet)
    _classify_rows(sheet)
    return sheet


def _stringify_tabular_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value)


def _fetch_google_sheet_metadata(spreadsheet_id: str) -> dict:
    cached = _GOOGLE_METADATA_CACHE.get(spreadsheet_id)
    if cached is not None:
        return cached
    service = _google_sheets_service()
    if service is None:
        raise RuntimeError(_google_private_sheet_auth_hint())
    metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id, includeGridData=False).execute()
    _GOOGLE_METADATA_CACHE[spreadsheet_id] = metadata
    return metadata


def _fetch_google_sheet_values(spreadsheet_id: str, sheet_name: str, gid: str = "") -> list[list[str]]:
    service = _google_sheets_service()
    if service is not None:
        response = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=_google_sheet_range(sheet_name))
            .execute()
        )
        return [[str(cell) for cell in row] for row in response.get("values", [])]
    return _fetch_public_google_sheet_csv(spreadsheet_id, sheet_name, gid=gid)


def _fetch_public_google_sheet_csv(spreadsheet_id: str, sheet_name: str, gid: str = "") -> list[list[str]]:
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq"
    params = {"tqx": "out:csv"}
    if gid:
        params["gid"] = gid
    else:
        params["sheet"] = sheet_name
    response = requests.get(url, params=params, timeout=20)
    if response.status_code != 200:
        raise RuntimeError(
            "无法读取 Google Sheets 内容。\n"
            f"{_google_private_sheet_auth_hint()}"
        )
    text = response.content.decode("utf-8-sig", errors="replace")
    if "<html" in text.lower() or "<!doctype" in text.lower():
        raise RuntimeError(
            "当前 Google Sheets 不能以公开 CSV 方式读取。\n"
            f"{_google_private_sheet_auth_hint()}"
        )
    return [row for row in csv.reader(io.StringIO(text))]


def _google_sheets_service():
    global _GOOGLE_SHEETS_SERVICE
    if _GOOGLE_SHEETS_SERVICE is not None:
        return _GOOGLE_SHEETS_SERVICE
    if build is None:
        return None
    credentials = _google_credentials([GOOGLE_SHEETS_READONLY_SCOPE])
    if credentials is None:
        return None
    _GOOGLE_SHEETS_SERVICE = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    return _GOOGLE_SHEETS_SERVICE


def _google_drive_service():
    global _GOOGLE_DRIVE_SERVICE
    if _GOOGLE_DRIVE_SERVICE is not None:
        return _GOOGLE_DRIVE_SERVICE
    if build is None:
        return None
    credentials = _google_credentials([GOOGLE_DRIVE_READONLY_SCOPE])
    if credentials is None:
        return None
    _GOOGLE_DRIVE_SERVICE = build("drive", "v3", credentials=credentials, cache_discovery=False)
    return _GOOGLE_DRIVE_SERVICE


def _google_credentials(scopes: list[str] | None = None):
    if _normalize_google_auth_mode(_GOOGLE_AUTH_MODE_OVERRIDE) == "oauth_user":
        return _load_google_oauth_credentials(scopes=scopes)
    if service_account is None:
        return None
    credentials_path = _google_credentials_path()
    if credentials_path is None:
        return None
    return service_account.Credentials.from_service_account_file(
        str(credentials_path),
        scopes=scopes or GOOGLE_READONLY_SCOPES,
    )


def _google_credentials_path() -> Path | None:
    env_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    candidates = [
        Path(_GOOGLE_SERVICE_ACCOUNT_OVERRIDE) if _GOOGLE_SERVICE_ACCOUNT_OVERRIDE else None,
        Path(env_path) if env_path else None,
        Path.cwd() / "google_service_account.json",
        Path.cwd() / "credentials" / "google_service_account.json",
    ]
    localappdata = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA")
    if localappdata:
        candidates.append(Path(localappdata) / "FenJiuBiHe" / "google_service_account.json")
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def _google_oauth_client_path() -> Path | None:
    auto_client = _auto_xlsx_to_xml_google_path("anotherProjectConfig", "credentials.json")
    candidates = [
        Path(_GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE) if _GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE else None,
        Path.cwd() / "google_oauth_client.json",
        Path.cwd() / "credentials" / "google_oauth_client.json",
        Path.cwd() / "credentials.json",
        auto_client,
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def _google_oauth_token_path() -> Path | None:
    auto_token = _auto_xlsx_to_xml_google_path("anotherProjectConfig", "token.json")
    candidates = [
        Path(_GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE) if _GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE else None,
        Path(default_google_oauth_token_path()),
        auto_token,
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def _load_google_oauth_credentials(scopes: list[str] | None = None):
    token_path = _google_oauth_token_path()
    if token_path is None:
        return None
    if UserCredentials is None:
        raise RuntimeError("当前环境缺少 google-auth，无法读取 Google 登录凭据。")
    try:
        credentials = UserCredentials.from_authorized_user_file(str(token_path), scopes=scopes)
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"读取 Google 登录凭据失败：{exc}") from exc
    if credentials.expired and credentials.refresh_token:
        if GoogleAuthRequest is None:
            raise RuntimeError("当前环境缺少 google.auth.transport.requests，无法刷新 Google 登录凭据。")
        try:
            credentials.refresh(GoogleAuthRequest())
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"刷新 Google 登录凭据失败，请重新登录：{exc}") from exc
        _save_google_oauth_token(credentials, token_path)
    if not credentials.valid:
        raise RuntimeError("当前 Google 登录已失效，请在“管理快捷配置”中重新执行 Google 登录。")
    return credentials


def _save_google_oauth_token(credentials, token_path: Path) -> None:
    token_path.parent.mkdir(parents=True, exist_ok=True)
    token_path.write_text(credentials.to_json(), encoding="utf-8")


def _google_private_sheet_auth_hint() -> str:
    if _normalize_google_auth_mode(_GOOGLE_AUTH_MODE_OVERRIDE) == "oauth_user":
        client_path = _GOOGLE_OAUTH_CLIENT_PATH_OVERRIDE or str(_google_oauth_client_path() or "")
        token_path = _GOOGLE_OAUTH_TOKEN_PATH_OVERRIDE or str(_google_oauth_token_path() or default_google_oauth_token_path())
        client_note = f"当前 OAuth 客户端：{client_path or '(未配置)'}"
        token_note = f"当前 OAuth token：{token_path or '(未配置)'}"
        return (
            "如果表格不是公开可访问，请在“管理快捷配置”中切到“个人登录”，"
            "可以直接选择现成的 OAuth token.json（例如 AutoXlsxtoXml 的 token.json）。\n"
            "只有在需要重新生成 token 时，才需要配置 OAuth 客户端 JSON 并点击“Google 登录”。\n"
            f"{client_note}\n{token_note}"
        )
    return (
        "如果表格不是公开可访问，请在“管理快捷配置”中提供 google_service_account.json，"
        "并把表格共享给 service account 邮箱。"
    )


def _normalize_google_auth_mode(value: str) -> str:
    text = str(value or "service_account").strip()
    return text if text in {"service_account", "oauth_user"} else "service_account"


def _run_google_oauth_login_manual(
    flow: InstalledAppFlow,
    manual_prompt_callback: Callable[[str], bool] | None = None,
):
    if oauth_flow_module is None:
        raise RuntimeError("当前环境缺少 google-auth-oauthlib，无法执行 Google 登录。")
    wsgi_app = oauth_flow_module._RedirectWSGIApp("Google 登录完成，现在可以回到程序。")
    wsgiref.simple_server.WSGIServer.allow_reuse_address = False
    local_server = wsgiref.simple_server.make_server(
        "localhost",
        0,
        wsgi_app,
        handler_class=oauth_flow_module._WSGIRequestHandler,
    )
    try:
        flow.redirect_uri = f"http://localhost:{local_server.server_port}/"
        auth_url, _ = flow.authorization_url()
        if manual_prompt_callback is None:
            raise RuntimeError(
                "无法自动打开浏览器进行 Google 授权。\n"
                "请手动访问以下链接完成授权：\n"
                f"{auth_url}"
            )
        if not manual_prompt_callback(auth_url):
            raise RuntimeError("已取消 Google 登录。")
        local_server.timeout = 300
        local_server.handle_request()
        try:
            authorization_response = wsgi_app.last_request_uri.replace("http", "https")
        except AttributeError as exc:
            raise RuntimeError("等待 Google 授权回调超时，请重试。") from exc
        flow.fetch_token(authorization_response=authorization_response)
        return flow.credentials
    finally:
        local_server.server_close()


def _reset_google_service_cache() -> None:
    global _GOOGLE_SHEETS_SERVICE
    global _GOOGLE_DRIVE_SERVICE
    _GOOGLE_SHEETS_SERVICE = None
    _GOOGLE_DRIVE_SERVICE = None
    _GOOGLE_METADATA_CACHE.clear()


def clear_sources_caches() -> None:
    _WORKBOOK_CACHE.clear()
    _REVISION_CACHE.clear()
    _GOOGLE_METADATA_CACHE.clear()


def _download_google_sheet_as_xlsx_bytes(spreadsheet_id: str) -> dict[str, object]:
    drive = _google_drive_service()
    if drive is None:
        raise RuntimeError(_google_private_sheet_auth_hint())
    if MediaIoBaseDownload is None:
        raise RuntimeError("当前环境缺少 google-api-client，无法下载 Google Sheets。")
    file_meta = (
        drive.files()
        .get(fileId=spreadsheet_id, fields="id,name,mimeType", supportsAllDrives=True)
        .execute()
    )
    request = drive.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    handle = io.BytesIO()
    downloader = MediaIoBaseDownload(handle, request, chunksize=1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    title = str(file_meta.get("name") or spreadsheet_id)
    file_name = title if title.lower().endswith(".xlsx") else f"{title}.xlsx"
    return {"title": title, "file_name": file_name, "content": handle.getvalue()}


def _google_sheet_range(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _looks_like_url(value: str) -> bool:
    lowered = value.lower()
    return lowered.startswith(("svn://", "svn:", "http://", "https://", "file://"))


def _auto_xlsx_to_xml_google_path(*parts: str) -> Path:
    return Path.cwd().parent.parent / "AutoXlsxtoXml" / Path(*parts)


def preferred_xml_roots(root: str | Path) -> list[str]:
    base = normalize_source_target(root)
    lowered = base.replace("\\", "/").lower().rstrip("/")
    if lowered.endswith("/xml_gy/core") or lowered.endswith("/xml_gy/normal"):
        return [base]
    if lowered.endswith("/xml_gy"):
        candidates = [
            join_source_target(base, "core"),
            join_source_target(base, "normal"),
        ]
        if not _looks_like_url(base):
            candidates.append(base)
        return _dedupe_preserve_order(candidates)
    candidates = [
        join_source_target(base, "xml_gy/core"),
        join_source_target(base, "xml_gy/normal"),
    ]
    if not _looks_like_url(base):
        candidates.append(base)
    return _dedupe_preserve_order(candidates)


def _list_svn_xml_files_from_target(target: str, recursive: bool = True, revision: str = "HEAD") -> list[SourceFileEntry]:
    args = ["svn", "list"]
    if recursive:
        args.append("-R")
    args.extend(["-r", str(revision or "HEAD").strip() or "HEAD"])
    args.append(target)
    output = run_svn_command(args)
    files: list[SourceFileEntry] = []
    for raw_line in output.splitlines():
        relative_path = raw_line.strip()
        if not relative_path or relative_path.endswith("/"):
            continue
        if not relative_path.lower().endswith(".xml"):
            continue
        files.append(
            SourceFileEntry(
                name=source_path_name(relative_path),
                path=join_source_target(target, relative_path),
            )
        )
    files.sort(key=lambda item: item.path.lower())
    return files


def _dedupe_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _dedupe_paths(values: list[Path]) -> list[Path]:
    result: list[Path] = []
    seen: set[str] = set()
    for value in values:
        normalized = str(value)
        if normalized in seen:
            continue
        seen.add(normalized)
        result.append(value)
    return result
