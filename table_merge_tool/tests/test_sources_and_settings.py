from __future__ import annotations
import csv
import io
from pathlib import Path

from openpyxl import Workbook

import table_merge_tool.sources as sources_module
from table_merge_tool.excel_xml import parse_workbook_bytes
from table_merge_tool.settings import (
    DEFAULT_SETTINGS,
    forget_quick_root,
    format_config_label,
    remember_config,
    remember_quick_root,
    replace_quick_roots,
    remember_root,
)
from table_merge_tool.sources import (
    WorkbookSource,
    build_snapshot_id,
    infer_source_kind,
    join_source_target,
    normalize_source_target,
    parse_google_sheet_target,
    preferred_xml_roots,
    set_google_auth_settings,
    set_google_service_account_path,
    source_relative_path,
    source_path_name,
    _google_credentials_path,
    _google_oauth_token_path,
    list_local_table_files,
    list_svn_xml_files,
    load_workbook_from_source,
)


SAMPLE_XML = b"""<?xml version="1.0" encoding="UTF-8"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
  <Worksheet ss:Name="Alpha">
    <Table>
      <Row><Cell><Data ss:Type="String">Name</Data></Cell></Row>
      <Row><Cell><Data ss:Type="String">name</Data></Cell></Row>
      <Row><Cell><Data ss:Type="String">A</Data></Cell></Row>
    </Table>
  </Worksheet>
  <Worksheet ss:Name="Beta">
    <Table>
      <Row><Cell><Data ss:Type="String">ID</Data></Cell></Row>
      <Row><Cell><Data ss:Type="String">id</Data></Cell></Row>
      <Row><Cell><Data ss:Type="String">1</Data></Cell></Row>
    </Table>
  </Worksheet>
</Workbook>
"""


def test_workbook_parses_sheet_names_lazily():
    workbook = parse_workbook_bytes(
        SAMPLE_XML,
        path=None,
        source_kind="local",
        source_label="sample.xml",
        snapshot_id="sample",
    )

    assert workbook.sheet_names == ["Alpha", "Beta"]
    assert workbook._sheet_cache == {}

    alpha = workbook.get_sheet("Alpha")

    assert alpha is not None
    assert alpha.name == "Alpha"
    assert "Alpha" in workbook._sheet_cache
    assert "Beta" not in workbook._sheet_cache


def test_xml_parser_trims_trailing_empty_styled_columns():
    payload = b"""<?xml version="1.0" encoding="UTF-8"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
  <Worksheet ss:Name="Wide">
    <Table ss:ExpandedColumnCount="500">
      <Row>
        <Cell><Data ss:Type="String">ID</Data></Cell>
        <Cell><Data ss:Type="String">Name</Data></Cell>
        <Cell ss:Index="500" ss:StyleID="sEmpty"><Data ss:Type="String"></Data></Cell>
      </Row>
      <Row>
        <Cell><Data ss:Type="String">id</Data></Cell>
        <Cell><Data ss:Type="String">name</Data></Cell>
        <Cell ss:Index="500" ss:StyleID="sEmpty"><Data ss:Type="String"></Data></Cell>
      </Row>
      <Row>
        <Cell><Data ss:Type="String">1</Data></Cell>
        <Cell><Data ss:Type="String">Alpha</Data></Cell>
        <Cell ss:Index="500" ss:StyleID="sEmpty"><Data ss:Type="String"></Data></Cell>
      </Row>
    </Table>
  </Worksheet>
</Workbook>
"""
    workbook = parse_workbook_bytes(
        payload,
        path=None,
        source_kind="local",
        source_label="wide.xml",
        snapshot_id="wide",
    )

    sheet = workbook.get_sheet("Wide")

    assert sheet is not None
    assert sheet.max_columns == 2
    assert len(sheet.logical_headers) == 2
    assert all(max((cell.column_index for cell in row.cells), default=0) <= 2 for row in sheet.rows)


def test_snapshot_id_uses_revision_for_svn_source():
    source = WorkbookSource(kind="svn", file_path="svn://repo/a.xml", revision="123")

    assert build_snapshot_id(source).endswith(":123")
    assert "svn://repo/a.xml" in build_snapshot_id(source)


def test_remember_root_keeps_recent_paths_unique():
    settings = dict(DEFAULT_SETTINGS)

    remember_root(settings, "C:/a")
    remember_root(settings, "C:/b")
    remember_root(settings, "C:/a")

    assert settings["recent_roots"] == ["C:/a", "C:/b"]


def test_default_settings_include_google_service_account_path():
    settings = dict(DEFAULT_SETTINGS)

    assert settings["google_service_account_path"] == ""
    assert settings["google_auth_mode"] == "service_account"
    assert settings["google_oauth_client_path"] == ""
    assert settings["google_oauth_token_path"] == ""
    assert settings["manual_key_fields"] == ""


def test_quick_roots_can_add_and_remove_entries():
    settings = dict(DEFAULT_SETTINGS)

    remember_quick_root(settings, "svn://repo/project/xml_gy/core", "主干 core")
    remember_quick_root(settings, "C:/xml", "本地 xml")
    remember_quick_root(settings, "svn://repo/project/xml_gy/core", "主干 core")
    forget_quick_root(settings, "C:/xml")

    assert settings["quick_roots"] == [{"name": "主干 core", "path": "svn://repo/project/xml_gy/core"}]


def test_replace_quick_roots_normalizes_legacy_entries():
    settings = dict(DEFAULT_SETTINGS)

    replace_quick_roots(
        settings,
        [
            "svn://repo/project/xml_gy/core",
            {"name": "normal", "path": "svn://repo/project/xml_gy/normal"},
        ],
    )

    assert settings["quick_roots"][0]["path"] == "svn://repo/project/xml_gy/core"
    assert settings["quick_roots"][0]["name"] == "xml_gy/core"
    assert settings["quick_roots"][1] == {"name": "normal", "path": "svn://repo/project/xml_gy/normal"}


def test_remember_config_keeps_latest_unique_snapshot():
    settings = dict(DEFAULT_SETTINGS)
    first = {
        "left_root": "C:/left",
        "right_root": "C:/right",
        "left_source": "local",
        "right_source": "svn",
        "left_file": "a.xml",
        "right_file": "b.xml",
        "left_revision": "WORKING",
        "right_revision": "123",
        "template_source": "right",
    }
    second = dict(first)
    second["right_revision"] = "124"

    remember_config(settings, first)
    remember_config(settings, second)
    remember_config(settings, first)

    assert len(settings["recent_configs"]) == 2
    assert settings["recent_configs"][0]["right_revision"] == "123"
    assert settings["recent_configs"][1]["right_revision"] == "124"


def test_format_config_label_contains_sources_and_files():
    label = format_config_label(
        {
            "left_source": "local",
            "right_source": "svn",
            "left_file": "left.xml",
            "right_file": "right.xml",
            "left_revision": "WORKING",
            "right_revision": "456",
            "template_source": "left",
            "manual_key_fields": "skillid,level",
        }
    )

    assert "LOCAL:left.xml" in label
    assert "SVN:right.xml@456" in label
    assert "左模板" in label
    assert "键:skillid,level" in label


def test_remember_config_preserves_manual_key_fields():
    settings = dict(DEFAULT_SETTINGS)

    remember_config(
        settings,
        {
            "left_root": "C:/left",
            "right_root": "C:/right",
            "left_file": "left.xml",
            "right_file": "right.xml",
            "manual_key_fields": "skillid,level",
        },
    )

    assert settings["recent_configs"][0]["manual_key_fields"] == "skillid,level"


def test_remember_config_preserves_strict_single_id_mode():
    settings = dict(DEFAULT_SETTINGS)

    remember_config(
        settings,
        {
            "left_root": "C:/left",
            "right_root": "C:/right",
            "left_file": "left.xml",
            "right_file": "right.xml",
            "manual_key_fields": "skillid",
            "strict_single_id_mode": True,
        },
    )

    assert settings["recent_configs"][0]["strict_single_id_mode"] is True
    assert "强ID:skillid" in format_config_label(settings["recent_configs"][0])


def test_remember_config_preserves_per_sheet_id_fields():
    settings = dict(DEFAULT_SETTINGS)

    remember_config(
        settings,
        {
            "left_root": "C:/left",
            "right_root": "C:/right",
            "left_file": "left.xml",
            "right_file": "right.xml",
            "sheet_key_fields": {"技能参数表": "skillid", "目标条件": "conditionid"},
        },
    )

    assert settings["recent_configs"][0]["sheet_key_fields"] == {
        "技能参数表": "skillid",
        "目标条件": "conditionid",
    }
    assert "检测ID:2" in format_config_label(settings["recent_configs"][0])


def test_join_source_target_and_name_support_svn_url():
    full_path = join_source_target("svn://repo/project/core", "tables/a.xml")

    assert full_path == "svn://repo/project/core/tables/a.xml"
    assert source_path_name(full_path) == "a.xml"


def test_join_source_target_keeps_absolute_svn_child():
    full_path = join_source_target("svn://repo/project/xml_gy", "svn://repo/project/xml_gy/core/a.xml")

    assert full_path == "svn://repo/project/xml_gy/core/a.xml"


def test_source_relative_path_supports_svn_subfolders():
    relative = source_relative_path("svn://repo/project/xml_gy", "svn://repo/project/xml_gy/core/skill_data.xml")

    assert relative == "core/skill_data.xml"


def test_normalize_source_target_supports_shorthand_svn_prefix():
    normalized = normalize_source_target(r"svn:192.168.6.20:13690\xdqd_xml\branch")

    assert normalized == "svn://192.168.6.20:13690/xdqd_xml/branch"


def test_infer_source_kind_distinguishes_svn_and_local():
    assert infer_source_kind("svn://repo/project") == "svn"
    assert infer_source_kind(r"svn:192.168.6.20:13690\xdqd_xml\branch") == "svn"
    assert infer_source_kind(r"C:\project\xml_gy\core") == "local"


def test_infer_source_kind_distinguishes_google_sheets_url_and_id():
    assert (
        infer_source_kind("https://docs.google.com/spreadsheets/d/1xFiaXtFs3NrR6-LzHKacFGzYNS6uhEA1bXfCFHkaaBI/edit?gid=1665856439")
        == "google_sheets"
    )
    assert infer_source_kind("1xFiaXtFs3NrR6-LzHKacFGzYNS6uhEA1bXfCFHkaaBI") == "google_sheets"


def test_parse_google_sheet_target_supports_url_and_gid():
    parsed = parse_google_sheet_target(
        "https://docs.google.com/spreadsheets/d/1xFiaXtFs3NrR6-LzHKacFGzYNS6uhEA1bXfCFHkaaBI/edit?gid=1665856439#gid=1665856439"
    )

    assert parsed == {
        "spreadsheet_id": "1xFiaXtFs3NrR6-LzHKacFGzYNS6uhEA1bXfCFHkaaBI",
        "gid": "1665856439",
    }


def test_google_service_account_override_path_takes_priority():
    credentials_path = Path(__file__).resolve()
    set_google_service_account_path(str(credentials_path))
    try:
        assert _google_credentials_path() == credentials_path
    finally:
        set_google_service_account_path("")


def test_google_oauth_token_override_path_takes_priority():
    token_path = Path(__file__).resolve()
    set_google_auth_settings(
        "oauth_user",
        service_account_path="",
        oauth_client_path="C:/fake/credentials.json",
        oauth_token_path=str(token_path),
    )
    try:
        assert _google_oauth_token_path() == token_path
    finally:
        set_google_auth_settings("service_account")


def test_preferred_xml_roots_prioritizes_core_and_normal():
    roots = preferred_xml_roots("svn://repo/project/branch")

    assert roots[:2] == [
        "svn://repo/project/branch/xml_gy/core",
        "svn://repo/project/branch/xml_gy/normal",
    ]


def test_preferred_xml_roots_handles_xml_gy_root_without_duplication():
    roots = preferred_xml_roots("svn://repo/project/branch/master/xml_gy")

    assert roots[:2] == [
        "svn://repo/project/branch/master/xml_gy/core",
        "svn://repo/project/branch/master/xml_gy/normal",
    ]
    assert "xml_gy/xml_gy" not in "\n".join(roots)


def test_list_svn_xml_files_uses_requested_revision(monkeypatch):
    calls: list[list[str]] = []

    def fake_run_svn_command(args: list[str]) -> str:
        calls.append(args)
        return "a.xml\nfolder/b.xml\nfolder/readme.txt\n"

    monkeypatch.setattr(sources_module, "preferred_xml_roots", lambda root: [root])
    monkeypatch.setattr(sources_module, "run_svn_command", fake_run_svn_command)

    files = list_svn_xml_files("svn://repo/project/xml_gy/normal", revision="12345")

    assert calls == [["svn", "list", "-R", "-r", "12345", "svn://repo/project/xml_gy/normal"]]
    assert [item.path for item in files] == [
        "svn://repo/project/xml_gy/normal/a.xml",
        "svn://repo/project/xml_gy/normal/folder/b.xml",
    ]


def test_list_local_table_files_includes_xml_xlsx_and_csv():
    test_dir = Path(__file__).resolve().parent / "_local_table_files_case"
    test_dir.mkdir(exist_ok=True)
    try:
        (test_dir / "a.xml").write_text("<Workbook/>", encoding="utf-8")
        (test_dir / "b.csv").write_text("id,name\n1,A\n", encoding="utf-8")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A1"] = "id"
        worksheet["B1"] = "name"
        workbook.save(test_dir / "c.xlsx")
        files = list_local_table_files(test_dir)
        assert [item.name for item in files] == ["a.xml", "b.csv", "c.xlsx"]
    finally:
        for child in test_dir.glob("*"):
            child.unlink()
        test_dir.rmdir()


def test_load_workbook_from_csv_source():
    test_path = Path(__file__).resolve().parent / "_workbook_case.csv"
    try:
        with test_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["id", "name"])
            writer.writerow(["1", "Alpha"])
        workbook = load_workbook_from_source(WorkbookSource(kind="local", file_path=str(test_path)))
        sheet = workbook.get_sheet(test_path.stem)
        assert workbook.sheet_names == [test_path.stem]
        assert sheet is not None
        assert sheet.max_columns == 2
        assert sheet.rows[1].value_at(1) == "1"
        assert workbook.source_meta.get("file_format") == "csv"
    finally:
        if test_path.exists():
            test_path.unlink()


def test_load_workbook_from_xlsx_source():
    test_path = Path(__file__).resolve().parent / "_workbook_case.xlsx"
    try:
        workbook_file = Workbook()
        worksheet = workbook_file.active
        worksheet.title = "Data"
        worksheet["A1"] = "id"
        worksheet["B1"] = "name"
        worksheet["A2"] = 1.0
        worksheet["B2"] = "Alpha"
        workbook_file.save(test_path)
        workbook = load_workbook_from_source(WorkbookSource(kind="local", file_path=str(test_path)))
        sheet = workbook.get_sheet("Data")
        assert workbook.sheet_names == ["Data"]
        assert sheet is not None
        assert sheet.max_columns == 2
        assert sheet.rows[1].value_at(1) == "1"
        assert sheet.rows[1].value_at(2) == "Alpha"
        assert workbook.source_meta.get("file_format") == "xlsx"
    finally:
        if test_path.exists():
            test_path.unlink()


def test_load_google_workbook_prefers_drive_exported_xlsx():
    workbook_file = Workbook()
    worksheet = workbook_file.active
    worksheet.title = "Skill_Data"
    worksheet["A1"] = "id"
    worksheet["B1"] = "name"
    worksheet["A2"] = 1.0
    worksheet["B2"] = "Alpha"
    buffer = io.BytesIO()
    workbook_file.save(buffer)
    payload = buffer.getvalue()

    original_download = sources_module._download_google_sheet_as_xlsx_bytes
    try:
        sources_module._download_google_sheet_as_xlsx_bytes = lambda spreadsheet_id: {
            "title": "Skill_Data",
            "file_name": "Skill_Data.xlsx",
            "content": payload,
        }
        workbook = load_workbook_from_source(
            WorkbookSource(
                kind="google_sheets",
                file_path="https://docs.google.com/spreadsheets/d/1xFiaXtFs3NrR6-LzHKacFGzYNS6uhEA1bXfCFHkaaBI/edit?gid=1665856439",
            )
        )
        sheet = workbook.get_sheet("Skill_Data")
        assert workbook.sheet_names == ["Skill_Data"]
        assert workbook.source_kind == "google_sheets"
        assert workbook.source_meta.get("google_load_mode") == "drive_export"
        assert sheet is not None
        assert sheet.rows[1].value_at(1) == "1"
        assert sheet.rows[1].value_at(2) == "Alpha"
    finally:
        sources_module._download_google_sheet_as_xlsx_bytes = original_download
